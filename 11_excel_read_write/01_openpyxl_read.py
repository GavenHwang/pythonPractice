# -*- coding:utf-8 -*-
import openpyxl

# 获取工作簿对象
wb = openpyxl.load_workbook('学生表.xlsx')

# 获取所有工作表名称
print(wb.sheetnames)

# 获取工作表对象,三种方法
sheet1 = wb.worksheets[0]
sheet2 = wb['Sheet1']
sheet3 = wb[wb.sheetnames[0]]
print(sheet1, sheet2, sheet3)

# 获取工作表名称
title = sheet1.title
print(title)

# 获取工作表总行数
rows = sheet1.max_row
# 获取工作表总列数
cols = sheet1.max_column
# 总行，总列
print(rows, cols)


# 获取某一单元格内容(行, 列)，例：2行1列，列表从1开始；也可以直接sheet1.cell(2, 1).value
cell = sheet1.cell(row=2, column=1).value
print(cell)

# 读取第一行的所有内容
row_list = []
for i in range(1, cols + 1):
    cell_value = sheet1.cell(row=1, column=i).value
    row_list.append(cell_value)

print(row_list)

# 读取第一列的所有内容
column_list = []
for i in range(1, rows + 1):
    cell_value = sheet1.cell(row=i, column=1).value
    column_list.append(cell_value)

print(column_list)

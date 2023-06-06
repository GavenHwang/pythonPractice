# -*- coding:utf-8 -*-
import openpyxl

# 创建一个新的工作簿对象
wb = openpyxl.Workbook()
# 获取工作表对象(sheet)
ws = wb.active
print(ws)
# 设置Sheet名称
ws.title = '学生表'

# 创建一个新sheet，可以指定名称，index表示新创建的工作簿放在第几个位置, index从0开始计数
ws_1 = wb.create_sheet(index=1, title='成绩表')
# 删除sheet表
del wb['成绩表']

# 写入单个单元格
ws['A1'] = '姓名'
ws['B1'] = '班级'
# 写入单个单元格(行，列，内容)
ws.cell(3, 4, '内容1')
# 写入多个单元格(追加模式，不会覆盖之前的，从有数据的下一行开始)
ws.append(['姓名', '班级', '年龄'])
ws.append(['王明', '三年级一班', '9岁'])

# 复制"学生表"，新sheet名称为"学生表 Copy"
ws_2 = wb.copy_worksheet(wb['学生表'])
# 修改"学生表 Copy"名称为"备份"
ws_2.title = '备份'

# 获取所有工作表名称
print(wb.sheetnames)

# 保存
wb.save('1.xlsx')

